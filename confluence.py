# -*- coding: utf-8 -*-
import requests
from requests.auth import HTTPBasicAuth
import argparse
import csv
import re
from io import BytesIO
from PIL import Image
import pytesseract
import os
import warnings
import time
import psutil
from datetime import datetime, timedelta
import logging
import json
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders

try:
    import boto3
except ImportError:
    boto3 = None

os.environ["TESSDATA_PREFIX"] = "/usr/local/share/tessdata/"
warnings.simplefilter('ignore', Image.DecompressionBombWarning)

try:
    import docx
except ImportError:
    docx = None

try:
    import fitz
except ImportError:
    fitz = None

try:
    import zipfile
    import tarfile
except ImportError:
    zipfile = tarfile = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

def setup_logging(log_file="confluence_scan.log"):
    """Configure logging to console and file."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )

def check_dependencies():
    """Check for required dependencies and return any errors."""
    errors = []
    if not docx:
        errors.append("python-docx is required for .docx support")
    if not fitz:
        errors.append("PyMuPDF is required for .pdf support")
    if not zipfile or not tarfile:
        errors.append("zipfile and tarfile are required for archive support")
    if not Image or not pytesseract:
        errors.append("PIL and pytesseract are required for image OCR")
    if not Workbook:
        errors.append("openpyxl is required for .xlsx support (pip install openpyxl)")
    return errors

def throttle_cpu(max_percent=50, interval=1):
    """Throttle CPU usage to prevent overloading."""
    proc = psutil.Process(os.getpid())
    while True:
        usage = proc.cpu_percent(interval=interval)
        if usage > max_percent:
            sleep_time = interval * (usage / max_percent - 1)
            time.sleep(min(sleep_time, 2))
        break

def safe_request(url, headers, auth=None, params=None):
    """Make a safe HTTP request with error handling."""
    try:
        resp = requests.get(url, headers=headers, auth=auth, params=params, timeout=30)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        logging.error(f"Request {url}: {e}")
        return {}

def normalize_base_url(base_url):
    """Normalize base URL for Atlassian Cloud."""
    base_url = base_url.rstrip('/')
    if 'atlassian.net' in base_url and '/wiki' not in base_url:
        return f"{base_url}/wiki"
    return base_url

def get_all_spaces(base_url, auth, headers):
    """Retrieve all spaces from Confluence."""
    all_spaces, start, limit = [], 0, 50
    while True:
        url = f"{base_url}/rest/api/space"
        data = safe_request(url, headers, auth, params={"start": start, "limit": limit})
        if not data:
            break
        all_spaces.extend(data.get("results", []))
        if "_links" in data and "next" in data["_links"]:
            start += limit
        else:
            break
    return all_spaces

def get_public_spaces(base_url, headers):
    """Retrieve only public spaces without authentication."""
    all_spaces, start, limit = [], 0, 50
    while True:
        url = f"{base_url}/rest/api/space"
        data = safe_request(url, headers, auth=None, params={"start": start, "limit": limit})
        if not data:
            break
        spaces = data.get("results", [])
        for space in spaces:
            space_key = space.get("key")
            test_url = f"{base_url}/rest/api/content"
            test_data = safe_request(
                test_url, headers, auth=None,
                params={"type": "page", "spaceKey": space_key, "limit": 1}
            )
            if test_data and test_data.get("results"):
                all_spaces.append(space)
                logging.info(f"Found public space: {space_key}")
        if "_links" in data and "next" in data["_links"]:
            start += limit
        else:
            break
    return all_spaces

def get_pages_in_space(base_url, auth, headers, space_key, modified_after=None, modified_before=None, created_in_year=None):
    """Retrieve pages in a space, applying date filters."""
    all_pages, start, limit = [], 0, 50
    while True:
        url = f"{base_url}/rest/api/content"
        params = {"type": "page", "spaceKey": space_key, "start": start, "limit": limit, "expand": "history,version"}
        data = safe_request(url, headers, auth, params=params)
        if not data:
            break
        pages = data.get("results", [])
        filtered_pages = [p for p in pages if filter_page(p, modified_after, modified_before, created_in_year)]
        for page in filtered_pages:
            page_id = page.get("id")
            full_data = safe_request(
                f"{base_url}/rest/api/content/{page_id}",
                headers, auth, params={"expand": "body.storage"}
            )
            if full_data:
                page["body"] = full_data.get("body", {})
                all_pages.append(page)
        if "_links" in data and "next" in data["_links"]:
            start += limit
        else:
            break
    return all_pages

def filter_page(page, modified_after, modified_before, created_in_year):
    """Filter pages based on modification and creation dates."""
    created_date_str = page.get("history", {}).get("createdDate")
    modified_date_str = page.get("version", {}).get("when")
    
    created_date = parse_iso_date(created_date_str) if created_date_str else None
    modified_date = parse_iso_date(modified_date_str) if modified_date_str else None
    
    if modified_after and modified_date and modified_date < modified_after:
        return False
    if modified_before and modified_date and modified_date > modified_before:
        return False
    if created_in_year and created_date and created_date.year != created_in_year:
        return False
    return True

def parse_iso_date(date_str):
    """Parse ISO date string to datetime."""
    try:
        return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
    except:
        return None

def get_last_editor_email(base_url, auth, headers, page_id):
    """Retrieve email of the last editor of a page."""
    try:
        url = f"{base_url}/rest/api/content/{page_id}/history"
        data = safe_request(url, headers, auth)
        if not data:
            return "?"
        
        last_updated = data.get("lastUpdated", {})
        by_user = last_updated.get("by", {})
        email = by_user.get("email", "?")
        return email if email else "?"
    except Exception as e:
        logging.error(f"Error fetching editor email for page {page_id}: {e}")
        return "?"

def get_attachments(base_url, auth, headers, page_id, max_age_delta=None):
    """Retrieve attachments for a given page."""
    try:
        url = f"{base_url}/rest/api/content/{page_id}/child/attachment"
        data = safe_request(url, headers, auth, params={"expand": "version"})
        if not data:
            return []
        
        attachments = data.get("results", [])
        if max_age_delta:
            cutoff = datetime.now() - max_age_delta
            filtered = []
            for att in attachments:
                when_str = att.get("version", {}).get("when", "")
                when_date = parse_iso_date(when_str)
                if when_date and when_date >= cutoff:
                    filtered.append(att)
            return filtered
        return attachments
    except Exception as e:
        logging.error(f"Error fetching attachments for page {page_id}: {e}")
        return []

def load_keywords(keywords_file):
    """Load keywords from a file."""
    if not keywords_file or not os.path.exists(keywords_file):
        return []
    with open(keywords_file, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip() and not line.startswith("#")]

def load_patterns(regex_file, single_regex):
    """Load regex patterns from file or single pattern."""
    patterns = []
    if regex_file and os.path.exists(regex_file):
        with open(regex_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and ":::" in line:
                    parts = line.split(":::")
                    if len(parts) >= 3:
                        name = parts[0].strip()
                        regex = parts[1].strip()
                        group_index = int(parts[2].strip())
                        patterns.append((name, regex, group_index))
    elif single_regex:
        patterns.append(("custom_regex", single_regex, 0))
    return patterns

def scan_text_for_keywords(text, keywords):
    """Scan text for keywords (case-insensitive)."""
    findings = []
    for keyword in keywords:
        if re.search(re.escape(keyword), text, re.IGNORECASE):
            findings.append(("keyword", keyword, keyword))
    return findings

def scan_text_for_secrets(text, patterns):
    """Scan text for secrets using regex patterns."""
    findings = []
    for pattern_name, pattern_regex, group_index in patterns:
        try:
            matches = re.finditer(pattern_regex, text, re.MULTILINE | re.IGNORECASE)
            for match in matches:
                if group_index < len(match.groups()) + 1:
                    secret_value = match.group(group_index) if group_index > 0 else match.group(0)
                    findings.append((pattern_name, secret_value))
        except re.error as e:
            logging.error(f"Regex error in pattern '{pattern_name}': {e}")
            continue
    return findings

def format_secret_value(value, max_length=20):
    """Format secret value for display (truncate if needed)."""
    value_str = str(value).strip()
    if len(value_str) > max_length:
        return value_str[:max_length] + "..."
    return value_str

def extract_text_from_attachment(base_url, auth, headers, attachment, max_size_bytes, keywords, patterns, scan_images_only, archive_support):
    """Extract text from various attachment types."""
    att_title = attachment.get("title", "?")
    download_url = f"{base_url}{attachment['_links'].get('download', '')}"
    
    file_size = attachment.get("extensions", {}).get("fileSize", 0)
    if max_size_bytes and file_size > max_size_bytes:
        logging.info(f"Skipping {att_title} (size: {file_size} bytes > max: {max_size_bytes})")
        return [], ""
    
    ext = os.path.splitext(att_title)[1][1:].lower()
    
    if scan_images_only and ext not in ["png", "jpg", "jpeg", "gif", "bmp", "tiff"]:
        return [], ext
    
    try:
        response = requests.get(download_url, auth=auth, headers=headers, timeout=30)
        response.raise_for_status()
        content = response.content
        
        text = ""
        
        # Text files (expanded list including scripts and code files)
        text_extensions = [
            # Basic text
            "txt", "log", "md", "markdown", "rst",
            # Config files
            "conf", "cfg", "ini", "properties", "env",
            # Data formats
            "json", "xml", "yaml", "yml", "toml", "csv", "tsv",
            # Scripts and code
            "py", "sh", "bash", "zsh", "fish", "bat", "cmd", "ps1",
            "js", "ts", "jsx", "tsx", "java", "c", "cpp", "h", "hpp",
            "cs", "go", "rs", "rb", "php", "pl", "r", "scala", "kt",
            "swift", "m", "mm", "dart", "lua", "groovy", "gradle",
            # Web
            "html", "htm", "css", "scss", "sass", "less",
            # Other
            "sql", "graphql", "proto", "dockerfile", "makefile",
            "terraform", "tf", "hcl", "jenkinsfile"
        ]
        
        if ext in text_extensions:
            text = content.decode("utf-8", errors="ignore")
        
        # DOCX
        elif ext == "docx" and docx:
            doc = docx.Document(BytesIO(content))
            text = "\n".join([para.text for para in doc.paragraphs])
        
        # PDF
        elif ext == "pdf" and fitz:
            pdf_doc = fitz.open(stream=content, filetype="pdf")
            text = "\n".join([page.get_text() for page in pdf_doc])
        
        # Images with OCR
        elif ext in ["png", "jpg", "jpeg", "gif", "bmp", "tiff"]:
            img = Image.open(BytesIO(content))
            text = pytesseract.image_to_string(img)
        
        # Archives (if enabled)
        elif archive_support and ext in ["zip", "tar", "gz", "tgz", "tar.gz"]:
            # Simple implementation: extract and scan text files
            if ext == "zip" and zipfile:
                with zipfile.ZipFile(BytesIO(content)) as zf:
                    for name in zf.namelist():
                        if name.endswith(('.txt', '.log', '.md', '.json', '.xml')):
                            text += zf.read(name).decode("utf-8", errors="ignore") + "\n"
        
        if not text:
            return [], ext
        
        # Scan for keywords and patterns
        findings = []
        if keywords:
            findings.extend(scan_text_for_keywords(text, keywords))
        if patterns:
            findings.extend([("regex", name, matched_text) for name, matched_text in scan_text_for_secrets(text, patterns)])
        
        return findings, ext
    
    except Exception as e:
        logging.error(f"Error extracting text from {att_title}: {e}")
        return [], ext

def parse_size(size_str):
    """Parse size string like '2mb' or '500kb' to bytes."""
    size_str = size_str.lower().strip()
    match = re.match(r"(\d+)(kb|mb|gb)?", size_str)
    if not match:
        return None
    num = int(match.group(1))
    unit = match.group(2) or "b"
    multipliers = {"b": 1, "kb": 1024, "mb": 1024**2, "gb": 1024**3}
    return num * multipliers.get(unit, 1)

def parse_date(date_str):
    """Parse date string in D.M.Y or D/M/Y format."""
    for sep in [".", "/"]:
        if sep in date_str:
            parts = date_str.split(sep)
            if len(parts) == 3:
                day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
                return datetime(year, month, day)
    raise ValueError(f"Invalid date format: {date_str}")

def parse_age(age_str):
    """Parse age string like '1d', '1w', '1m', '1y' to timedelta."""
    age_str = age_str.lower().strip()
    match = re.match(r"(\d+)([dwmy])", age_str)
    if not match:
        return None
    num = int(match.group(1))
    unit = match.group(2)
    if unit == "d":
        return timedelta(days=num)
    elif unit == "w":
        return timedelta(weeks=num)
    elif unit == "m":
        return timedelta(days=num * 30)
    elif unit == "y":
        return timedelta(days=num * 365)
    return None

def load_list_from_arg(arg):
    """Load list from comma-separated string or file."""
    if not arg:
        return set()
    if os.path.isfile(arg):
        with open(arg, "r", encoding="utf-8") as f:
            return set([line.strip() for line in f if line.strip()])
    return set([k.strip() for k in arg.split(",")])

def validate_arguments(args):
    """Validate command-line arguments."""
    errors = []
    
    if not args.base_url:
        errors.append("--base-url is required")
    
    if not args.public_only:
        if not args.username:
            errors.append("--username is required (or use --public-only)")
        if not args.token:
            errors.append("--token is required (or use --public-only)")
    
    if not args.keywords and not args.regex_file and not args.regex:
        errors.append("At least one of --keywords, --regex, or --regex-file is required")
    
    if args.include_attachments:
        if args.filetype and args.exclude_filetype:
            errors.append("Cannot use both --filetype and --exclude-filetype")
    
    if args.email_sender and not args.email_recipient:
        errors.append("--email-recipient is required when using --email-sender")
    
    if args.email_recipient and not args.email_sender:
        errors.append("--email-sender is required when using --email-recipient")
    
    # Validate alert flag
    if args.alert and not args.email_sender:
        errors.append("--email-sender is required when using --alert")
    
    return errors

def load_config(config_file):
    """Load configuration from JSON file."""
    if not config_file or not os.path.exists(config_file):
        return {}
    try:
        with open(config_file, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        logging.error(f"Error loading config file: {e}")
        return {}

def create_xlsx_report(csv_file, xlsx_file):
    """Convert CSV to formatted XLSX report with improved formatting."""
    if not Workbook:
        logging.warning("openpyxl not installed, skipping XLSX creation")
        return False
    
    try:
        # Read CSV
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
        
        if len(rows) < 2:
            logging.warning("No data to create XLSX report")
            return False
        
        # Remove dummy row if present
        if len(rows) > 1 and "DUMMY_ROW_DELETE_ME" in rows[1][0]:
            rows.pop(1)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Confluence Secrets"
        
        # Define styles
        header_fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write data
        for row_idx, row_data in enumerate(rows, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                
                # Header styling (first row)
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Secret value in red (last column or "Matched Value")
                    if col_idx == len(row_data) or (row_idx == 1 and 'Matched Value' in str(value)):
                        if row_idx > 1:  # Don't color header
                            cell.font = Font(color='DC143C', bold=True)
                    
                    # Make URLs clickable
                    if 'http' in str(value).lower() and '://' in str(value):
                        cell.hyperlink = value
                        cell.font = Font(color='0563C1', underline='single')
        
        # Auto-adjust column widths based on content
        for col_idx in range(1, len(rows[0]) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            # Check header
            header_value = ws.cell(row=1, column=col_idx).value
            if header_value:
                max_length = len(str(header_value))
            
            # Check first 100 rows for performance
            for row_idx in range(2, min(len(rows) + 1, 102)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    # For URLs, limit the length calculation
                    if 'http' in str(cell_value).lower():
                        max_length = max(max_length, min(len(str(cell_value)), 60))
                    else:
                        max_length = max(max_length, min(len(str(cell_value)), 50))
            
            # Set column width with reasonable limits
            adjusted_width = min(max_length + 2, 70)
            if adjusted_width < 15:
                adjusted_width = 15
            
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set header row height
        ws.row_dimensions[1].height = 30
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Save workbook
        wb.save(xlsx_file)
        logging.info(f"XLSX report created: {xlsx_file}")
        return True
        
    except Exception as e:
        logging.error(f"Error creating XLSX report: {e}")
        return False

def create_author_report(author_findings, author_email, filename):
    """
    Create personalized Excel report for a specific author
    
    Args:
        author_findings: List of findings for this author
        author_email: Author's email address
        filename: Output filename for the report
    
    Returns:
        Filename of created report
    """
    if not Workbook:
        logging.error("openpyxl not installed, cannot create author report")
        return None
    
    try:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Your Confluence Secrets"
        
        # Define border
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Check if any findings have attachment info
        has_attachments = any(f.get('is_attachment', False) for f in author_findings)
        
        # Row 1: Main title with red background
        if has_attachments:
            headers = ['Space Name', 'Page Title', 'File Name', 'File Type', 'File URL', 'Page URL', 'Matched Keyword', 'Finding Type', 'Matched Value']
        else:
            headers = ['Space Name', 'Page Title', 'Page URL', 'Matched Keyword', 'Finding Type', 'Matched Value']
        
        last_col = get_column_letter(len(headers))
        
        sheet.merge_cells(f'A1:{last_col}1')
        title_cell = sheet.cell(row=1, column=1)
        title_cell.value = "SECURITY ALERT: Exposed Secrets Found in Your Confluence Pages"
        title_cell.font = Font(size=14, bold=True, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
        sheet.row_dimensions[1].height = 25
        
        # Row 2: Warning message with yellow background
        sheet.merge_cells(f'A2:{last_col}2')
        warning_cell = sheet.cell(row=2, column=1)
        warning_cell.value = "The following secrets were detected in pages you edited. Please take immediate action."
        warning_cell.font = Font(size=10, italic=True)
        warning_cell.alignment = Alignment(horizontal='center', vertical='center')
        warning_cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        sheet.row_dimensions[2].height = 20
        
        # Row 3: Empty row for spacing
        sheet.row_dimensions[3].height = 5
        
        # Row 4: Column headers with red background
        header_fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        sheet.row_dimensions[4].height = 30
        
        # Add findings (starting from row 5)
        for row_num, finding in enumerate(author_findings, 5):
            col = 1
            
            # Space Name
            cell = sheet.cell(row=row_num, column=col)
            cell.value = finding.get('space_name', 'Unknown')
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            col += 1
            
            # Page Title
            cell = sheet.cell(row=row_num, column=col)
            cell.value = finding.get('page_title', 'Unknown')
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            col += 1
            
            if has_attachments:
                # File Name (only if this finding has attachment info)
                cell = sheet.cell(row=row_num, column=col)
                if finding.get('is_attachment', False):
                    cell.value = finding.get('file_title', 'N/A')
                else:
                    cell.value = '(Page Content)'
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                col += 1
                
                # File Type
                cell = sheet.cell(row=row_num, column=col)
                if finding.get('is_attachment', False):
                    cell.value = finding.get('file_extension', 'N/A')
                else:
                    cell.value = 'N/A'
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                col += 1
                
                # File URL
                cell = sheet.cell(row=row_num, column=col)
                if finding.get('is_attachment', False):
                    file_url = finding.get('file_url', '')
                    cell.value = file_url
                    if file_url:
                        cell.hyperlink = file_url
                        cell.font = Font(color='0563C1', underline='single')
                else:
                    cell.value = 'N/A'
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                col += 1
            
            # Page URL as hyperlink
            url_cell = sheet.cell(row=row_num, column=col)
            page_url = finding.get('page_url', '')
            url_cell.value = page_url
            if page_url:
                url_cell.hyperlink = page_url
                url_cell.font = Font(color='0563C1', underline='single')
            url_cell.border = border
            url_cell.alignment = Alignment(wrap_text=True, vertical='top')
            col += 1
            
            # Matched Keyword
            cell = sheet.cell(row=row_num, column=col)
            cell.value = finding.get('keyword', 'Unknown')
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            col += 1
            
            # Finding Type
            cell = sheet.cell(row=row_num, column=col)
            cell.value = finding.get('finding_type', 'unknown')
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            col += 1
            
            # Secret value - red bold font
            secret_cell = sheet.cell(row=row_num, column=col)
            secret_cell.value = finding.get('matched_value', '')
            secret_cell.font = Font(color='DC143C', bold=True, size=10)
            secret_cell.border = border
            secret_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Auto-adjust column widths based on content
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            
            # Calculate max length for this column
            max_length = len(header)
            for row_num in range(5, 5 + len(author_findings)):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    # For URLs and long text, limit the calculation
                    if 'URL' in header or 'http' in str(cell_value).lower():
                        max_length = max(max_length, min(len(str(cell_value)), 60))
                    else:
                        max_length = max(max_length, min(len(str(cell_value)), 50))
            
            # Set column width with reasonable limits
            adjusted_width = min(max_length + 2, 70)
            if adjusted_width < 15:
                adjusted_width = 15
            
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze top rows (header and title)
        sheet.freeze_panes = "A5"
        
        wb.save(filename)
        return filename
    
    except Exception as e:
        logging.error(f"Error creating author report for {author_email}: {e}")
        return None


def send_author_alert(author_email, author_name, author_findings, aws_region='eu-central-1', 
                     sender_email=None, security_contact='security@company.com', 
                     security_wiki=None, has_attachments=False):
    """
    Send personalized alert email to page editor
    
    Args:
        author_email: Editor's email address
        author_name: Editor's display name (can be email if name not available)
        author_findings: List of secrets found in pages edited by this person
        aws_region: AWS region for SES
        sender_email: Sender email address
        security_contact: Security team contact email
        security_wiki: Security documentation URL (optional)
        has_attachments: Boolean indicating if scan included attachments
    
    Returns:
        Boolean indicating success
    """
    if not boto3:
        logging.error(f"      ❌ boto3 not available for {author_email}")
        return False
    
    if not sender_email:
        logging.error(f"      ❌ Sender email not specified for {author_email}")
        return False
    
    try:
        # Create SES client
        ses_client = boto3.client("ses", region_name=aws_region)
        
        # Create author-specific report
        temp_filename = f"temp_author_{author_email.replace('@', '_at_').replace('.', '_')}.xlsx"
        create_author_report(author_findings, author_email, temp_filename)
        
        # Prepare subject
        secret_count = len(author_findings)
        page_count = len(set(f['page_url'] for f in author_findings))
        subject = f"SECURITY ALERT: {secret_count} Secret{'s' if secret_count != 1 else ''} Found in Your Confluence Page{'s' if page_count != 1 else ''}"
        
        # Prepare email body
        body_text = f"""Hello {author_name if author_name and author_name != author_email else 'there'},

We have detected {secret_count} exposed secret{'s' if secret_count != 1 else ''} in {page_count} Confluence page{'s' if page_count != 1 else ''} that you last edited.

AFFECTED PAGES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
        
        # List affected pages
        for page_url in sorted(set(f['page_url'] for f in author_findings)):
            page_findings = [f for f in author_findings if f['page_url'] == page_url]
            page_title = page_findings[0].get('page_title', 'Unknown')
            body_text += f"\n• {page_title}: {len(page_findings)} secret{'s' if len(page_findings) != 1 else ''} found\n"
            body_text += f"  {page_url}\n"
        
        body_text += f"""
IMMEDIATE ACTION REQUIRED
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

1. Review the attached Excel report for complete details
2. Rotate/revoke ALL exposed credentials immediately
3. Remove secrets from the affected Confluence pages{' and attachments' if has_attachments else ''}
4. Update applications with new credentials
5. Never store secrets in Confluence pages, comments, or attachments

The detailed report is attached to this email. Please address these issues immediately.

If you have any questions, contact your security team: {security_contact}
"""
        
        if security_wiki:
            body_text += f"\nSecurity documentation: {security_wiki}\n"
        
        body_text += """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
This is an automated security alert from Confluence Secrets Scanner.
"""
        
        # Create email message
        msg = MIMEMultipart()
        msg["From"] = sender_email
        msg["To"] = author_email
        msg["Subject"] = subject
        
        # Attach body
        msg.attach(MIMEText(body_text, "plain"))
        
        # Determine attachment filename based on scan mode
        if has_attachments:
            attachment_filename = "your_confluence_secrets_in_files.xlsx"
        else:
            attachment_filename = "your_confluence_secrets.xlsx"
        
        # Attach Excel report with appropriate name
        if os.path.exists(temp_filename):
            with open(temp_filename, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {attachment_filename}",
            )
            msg.attach(part)
        
        # Send email
        response = ses_client.send_raw_email(
            Source=sender_email,
            Destinations=[author_email],
            RawMessage={"Data": msg.as_string()}
        )
        
        # Clean up temp file
        if os.path.exists(temp_filename):
            os.remove(temp_filename)
        
        logging.info(f"      ✅ Alert sent to {author_name} ({author_email})")
        return True
        
    except Exception as e:
        logging.error(f"      ❌ Failed to send alert to {author_email}: {e}")
        # Clean up temp file on error
        if os.path.exists(temp_filename):
            try:
                os.remove(temp_filename)
            except:
                pass
        return False


def generate_email_summary(total_secrets, total_pages, total_files, spaces_scanned, duration_str, include_attachments, affected_spaces=None, affected_pages=None):
    """Generate email summary in the requested format"""
    timestamp = datetime.now().strftime("%d.%m.%Y %H:%M")
    
    # Use actual affected counts if provided, otherwise use scanned counts
    actual_spaces = affected_spaces if affected_spaces is not None else spaces_scanned
    actual_pages = affected_pages if affected_pages is not None else total_pages
    
    # Build email body
    body = f"""Confluence Secrets Scanner Report - {timestamp}

Summary Statistics:
* Total Secrets Found: {total_secrets}
* Affected Spaces: {actual_spaces}
* Affected Pages: {actual_pages}"""
    
    if include_attachments:
        body += f"\n* Files Scanned: {total_files}"
    
    body += f"""

ACTION REQUIRED:
1. Review the attached report immediately
2. Rotate/revoke exposed credentials
3. Implement proper secrets management

The detailed report is attached as XLSX file.

---
This is an automated report generated by Confluence Secrets Scanner.
"""
    
    return body

def count_affected_from_csv(csv_file):
    """Count unique affected spaces and pages from CSV results"""
    try:
        if not os.path.exists(csv_file):
            return None, None
        
        unique_spaces = set()
        unique_pages = set()
        
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader)  # Skip header
            
            for row in reader:
                if len(row) >= 3:  # Need at least space_name, title, page_url
                    space_name = row[0].strip()
                    page_url = row[2].strip() if len(row) > 2 else ""
                    
                    if space_name:
                        unique_spaces.add(space_name)
                    if page_url:
                        unique_pages.add(page_url)
        
        return len(unique_spaces), len(unique_pages)
    
    except Exception as e:
        logging.warning(f"Could not count affected spaces/pages from CSV: {e}")
        return None, None

def send_email_with_attachment(subject, body_text, sender, recipient, aws_region, attachment_path):
    """Send email via AWS SES with XLSX attachment
    
    Args:
        recipient: String with single email or comma-separated emails (e.g., "user1@example.com,user2@example.com")
    """
    if not boto3:
        logging.error("boto3 not installed. Install with: pip install boto3")
        return False
    
    try:
        # Parse recipients - support both single email and comma-separated list
        if isinstance(recipient, str):
            recipients = [email.strip() for email in recipient.split(',') if email.strip()]
        else:
            recipients = [recipient]
        
        if not recipients:
            logging.error("No valid recipients specified")
            return False
        
        # Create SES client
        ses_client = boto3.client("ses", region_name=aws_region)
        
        # Create email message
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = ", ".join(recipients)  # Join all recipients for the header
        msg["Subject"] = subject
        
        # Attach body
        msg.attach(MIMEText(body_text, "plain", "utf-8"))
        
        # Attach XLSX file
        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, "rb") as f:
                attachment = MIMEApplication(f.read(), _subtype="xlsx")
                attachment.add_header(
                    "Content-Disposition", 
                    "attachment", 
                    filename=os.path.basename(attachment_path)
                )
                msg.attach(attachment)
            logging.info(f"Attached file: {attachment_path}")
        
        # Send email to all recipients
        response = ses_client.send_raw_email(
            Source=sender,
            Destinations=recipients,  # AWS SES expects a list
            RawMessage={"Data": msg.as_string()}
        )
        
        logging.info(f"✅ Email sent successfully to {len(recipients)} recipient(s): {', '.join(recipients)}")
        logging.info(f"   MessageId: {response['MessageId']}")
        return True
        
    except Exception as e:
        logging.error(f"❌ Error sending email: {e}")
        return False

def validate_csv_fields(space_name, title, page_url, keyword, finding_type, formatted_value, email):
    """
    Validate and sanitize all fields before writing to CSV.
    Ensures all fields are non-empty strings to prevent column shifting.
    """
    # Validate and provide defaults for each field
    safe_space_name = str(space_name).strip() if space_name else "Unknown Space"
    safe_title = str(title).strip() if title and str(title).strip() != "?" else "Untitled Page"
    safe_page_url = str(page_url).strip() if page_url else "No URL"
    safe_keyword = str(keyword).strip() if keyword else "Unknown Keyword"
    safe_finding_type = str(finding_type).strip() if finding_type else "unknown"
    safe_formatted_value = str(formatted_value).strip() if formatted_value else "No Value"
    safe_email = str(email).strip() if email and str(email).strip() != "?" else "unknown@unknown.com"
    
    # Log warning if any field was replaced with default
    if not space_name or not str(space_name).strip():
        logging.warning(f"Empty space_name, using default: {safe_space_name}")
    if not title or str(title).strip() == "?":
        logging.warning(f"Empty or missing title for URL {safe_page_url}, using default: {safe_title}")
    if not email or str(email).strip() == "?":
        logging.warning(f"Empty or missing email for {safe_title}, using default: {safe_email}")
    
    return (safe_space_name, safe_title, safe_page_url, safe_keyword, 
            safe_finding_type, safe_formatted_value, safe_email)

def validate_attachment_fields(space_name, page_title, file_title, ext, file_url, page_url, keyword, finding_type, formatted_value, email):
    """
    Validate and sanitize attachment fields before writing to CSV.
    Ensures all 10 fields are non-empty strings to prevent column shifting.
    """
    safe_space_name = str(space_name).strip() if space_name else "Unknown Space"
    safe_page_title = str(page_title).strip() if page_title and str(page_title).strip() != "?" else "Untitled Page"
    safe_file_title = str(file_title).strip() if file_title and str(file_title).strip() != "?" else "Unknown File"
    safe_ext = str(ext).strip() if ext else "unknown"
    safe_file_url = str(file_url).strip() if file_url else "No URL"
    safe_page_url = str(page_url).strip() if page_url else "No URL"
    safe_keyword = str(keyword).strip() if keyword else "Unknown Keyword"
    safe_finding_type = str(finding_type).strip() if finding_type else "unknown"
    safe_formatted_value = str(formatted_value).strip() if formatted_value else "No Value"
    safe_email = str(email).strip() if email and str(email).strip() != "?" else "unknown@unknown.com"
    
    # Log warnings
    if not page_title or str(page_title).strip() == "?":
        logging.warning(f"Empty page title for attachment {safe_file_title}")
    if not file_title or str(file_title).strip() == "?":
        logging.warning(f"Empty file title in page {safe_page_title}")
    
    return (safe_space_name, safe_page_title, safe_file_title, safe_ext, safe_file_url,
            safe_page_url, safe_keyword, safe_finding_type, safe_formatted_value, safe_email)

def process_space(base_url, auth, headers, space, keywords, patterns, writer, csvfile, include_attachments, allowed_types, excluded_types, max_size_bytes, mod_after, mod_before, created_year, no_duplicates, max_age_delta, secret_max_length, scan_images_only, archive_support, findings_set, debug_limit=None, current_total=0):
    """Process a single Confluence space."""
    space_key = space.get("key", "?")
    space_name = space.get("name", space_key).strip()
    logging.info(f"Processing space: {space_key} - {space_name}")
    pages = get_pages_in_space(base_url, auth, headers, space_key, mod_after, mod_before, created_year)
    
    total_files_in_space = 0
    total_secrets_in_space = 0
    total_pages_in_space = len(pages)
    
    # Don't reopen file - use existing writer
    for page in pages:
        secrets_found, files_scanned = process_page(base_url, auth, headers, page, space_name, space_key, writer, csvfile, keywords, patterns, include_attachments, allowed_types, excluded_types, max_size_bytes, no_duplicates, max_age_delta, secret_max_length,
            scan_images_only, archive_support, findings_set
        )
        total_secrets_in_space += secrets_found
        total_files_in_space += files_scanned
        
        # Debug mode: stop if reached limit
        if debug_limit and (current_total + total_secrets_in_space) >= debug_limit:
            logging.warning(f"🔍 DEBUG MODE: Reached limit in space '{space_name}'. Stopping...")
            break
    
    summary = f"[SUMMARY] Space '{space_name}': {total_pages_in_space} pages scanned, {total_secrets_in_space} secrets found"
    if include_attachments:
        summary += f", {total_files_in_space} files scanned"
    logging.info(summary)
    
    return total_secrets_in_space, total_files_in_space, total_pages_in_space

def process_page(base_url, auth, headers, page, space_name, space_key, writer, csvfile, keywords, patterns, include_attachments, allowed_types, excluded_types, max_size_bytes, no_duplicates, max_age_delta, secret_max_length, scan_images_only, archive_support, findings_set):
    """Process a single Confluence page."""
    title = page.get("title", "?")
    page_id = page.get("id", "?")
    page_url = f"{base_url}/spaces/{space_key}/pages/{page_id}"
    total_files_in_page = 0
    secrets_found = 0
    
    if include_attachments:
        attachments = get_attachments(base_url, auth, headers, page_id, max_age_delta)
        total_files_in_page = len(attachments)
        filtered_attachments = [
            att for att in attachments
            if (allowed_types is None or os.path.splitext(att.get("title", ""))[1][1:].lower() in allowed_types)
            and os.path.splitext(att.get("title", ""))[1][1:].lower() not in excluded_types
        ]
        for att in filtered_attachments:
            file_findings, ext = extract_text_from_attachment(
                base_url, auth, headers, att, max_size_bytes, keywords, patterns,
                scan_images_only, archive_support
            )
            if not file_findings:
                continue
            for finding_type, keyword, matched_text in file_findings:
                formatted_value = format_secret_value(matched_text, max_length=secret_max_length)
                file_url = f"{base_url}{att['_links'].get('download', '')}"
                email = get_last_editor_email(base_url, auth, headers, page_id)
                
                # Validate all attachment fields before writing to CSV
                finding_tuple = validate_attachment_fields(
                    space_name, title, att.get("title", "?"), ext, 
                    file_url, page_url, keyword, finding_type, 
                    formatted_value, email
                )
                
                if no_duplicates and finding_tuple in findings_set:
                    continue
                if no_duplicates:
                    findings_set.add(finding_tuple)
                logging.info(f"[FOUND] Page: {title}, File: {att.get('title', '?')}, Keyword: {keyword}, Type: {finding_type}, Value: {formatted_value}")
                writer.writerow(list(finding_tuple))
                csvfile.flush()
                secrets_found += 1
    else:
        body = page.get("body", {}).get("storage", {}).get("value", "")
        findings = []
        if keywords:
            findings.extend(scan_text_for_keywords(body, keywords))
        if patterns:
            findings.extend([("regex", name, matched_text) for name, matched_text in scan_text_for_secrets(body, patterns)])
        for finding_type, keyword, matched_text in findings:
            formatted_value = format_secret_value(matched_text, max_length=secret_max_length)
            email = get_last_editor_email(base_url, auth, headers, page_id)
            
            # Validate all fields before writing to CSV
            finding_tuple = validate_csv_fields(
                space_name, title, page_url, keyword, 
                finding_type, formatted_value, email
            )
            
            if no_duplicates and finding_tuple in findings_set:
                continue
            if no_duplicates:
                findings_set.add(finding_tuple)
            logging.info(f"[FOUND] Page: {title}, Keyword: {keyword}, Type: {finding_type}, Value: {formatted_value}")
            writer.writerow(list(finding_tuple))
            csvfile.flush()
            secrets_found += 1
    
    return secrets_found, total_files_in_page

def main(base_url, username, token, keywords_file, regex_file, single_regex, output_file, 
         include_attachments, filetypes, exclude_filetypes, max_size, public_only,
         space_keys, exclude_space_keys, modified_after, modified_before, created_in_year,
         no_duplicates, resume_from, max_attachment_age, scan_images_only, archive_support, 
         config, email_sender, email_recipient, aws_region, secret_max_length, debug_mode,
         alert, security_contact, security_wiki):
    """Main function to scan Confluence for keywords and secrets."""
    start_time = time.time()
    setup_logging()
    
    # Load config file and override unset arguments
    config_dict = load_config(config)
    
    # Check dependencies
    dep_errors = check_dependencies()
    if dep_errors and (include_attachments or scan_images_only or archive_support):
        for error in dep_errors:
            logging.error(error)
        exit(1)
    
    base_url = normalize_base_url(base_url)
    logging.info(f"Using base URL: {base_url}")
    
    # Debug mode warning
    if debug_mode:
        logging.warning("=" * 80)
        logging.warning("🔍 DEBUG MODE ENABLED - Script will stop after first 5 findings!")
        logging.warning("=" * 80)
    
    if public_only:
        auth = None
        logging.info("Running in PUBLIC-ONLY mode (no authentication)")
    else:
        auth = HTTPBasicAuth(username, token)
        logging.info("Running in AUTHENTICATED mode")
    
    headers = {"Accept": "application/json"}
    keywords = load_keywords(keywords_file) if keywords_file else []
    patterns = load_patterns(regex_file, single_regex)
    allowed_types = set([t.lower() for t in filetypes.split(",")]) if filetypes else None
    excluded_types = set([t.lower() for t in exclude_filetypes.split(",")]) if exclude_filetypes else set()
    max_size_bytes = parse_size(max_size) if max_size else None
    space_keys_set = load_list_from_arg(space_keys)
    exclude_space_keys_set = load_list_from_arg(exclude_space_keys)
    try:
        mod_after = parse_date(modified_after) if modified_after else None
    except ValueError as e:
        logging.error(e)
        mod_after = None
    try:
        mod_before = parse_date(modified_before) if modified_before else None
    except ValueError as e:
        logging.error(e)
        mod_before = None
    created_year = int(created_in_year) if created_in_year else None
    max_age_delta = parse_age(max_attachment_age) if max_attachment_age else None

    if not keywords and not patterns:
        logging.error("No search criteria provided. Specify --keywords, --regex, or --regex-file")
        return

    if keywords:
        logging.info(f"Loaded keywords: {len(keywords)}")
    if patterns:
        logging.info(f"Loaded regex patterns: {len(patterns)}")
    
    if public_only:
        logging.info("Searching for public spaces...")
        spaces = get_public_spaces(base_url, headers)
    else:
        spaces = get_all_spaces(base_url, auth, headers)
    
    if space_keys_set:
        spaces = [s for s in spaces if s.get("key") in space_keys_set]
    spaces = [s for s in spaces if s.get("key") not in exclude_space_keys_set]
    spaces.sort(key=lambda s: s.get("key", ""))
    if resume_from:
        spaces = [s for s in spaces if s.get("key") >= resume_from]
    
    logging.info(f"Found/filtered spaces: {len(spaces)}")

    findings_set = set() if no_duplicates else None
    
    total_secrets = 0
    total_files = 0
    total_pages = 0
    
    # Debug mode: stop after 5 findings
    debug_limit = 5 if debug_mode else None
    
    # Store all findings for author alerts
    all_findings_data = []
    
    with open(output_file, "w", encoding="utf-8", newline="") as csvfile:
        writer = csv.writer(csvfile)
        if include_attachments:
            writer.writerow([
                "Space Name", "Page Title", "File Title", "File Extension", "File URL", 
                "Page URL", "Matched Keyword", "Finding Type", "Matched Value", "Last Editor Email"
            ])
            # Dummy row to absorb the first-row bug
            writer.writerow([
                "DUMMY_ROW_DELETE_ME", "DUMMY", "DUMMY", "txt", "http://dummy.com/file", 
                "http://dummy.com/page", "DUMMY_KEYWORD", "dummy", "DUMMY_VALUE", "dummy@dummy.com"
            ])
        else:
            writer.writerow([
                "Space Name", "Page Title", "Page URL", "Matched Keyword", 
                "Finding Type", "Matched Value", "Last Editor Email"
            ])
            # Dummy row to absorb the first-row bug
            writer.writerow([
                "DUMMY_ROW_DELETE_ME", "DUMMY", "http://dummy.com/page", "DUMMY_KEYWORD", 
                "dummy", "DUMMY_VALUE", "dummy@dummy.com"
            ])
        
        logging.info("🔧 Added dummy row after headers to prevent first-row bug")

        for space in spaces:
            secrets_in_space, files_in_space, pages_in_space = process_space(
                base_url, auth, headers, space, keywords, patterns, writer, csvfile,
                include_attachments, allowed_types, excluded_types, max_size_bytes, 
                mod_after, mod_before, created_year, no_duplicates, max_age_delta, 
                secret_max_length, scan_images_only, archive_support, findings_set, 
                debug_limit, total_secrets
            )
            total_secrets += secrets_in_space
            total_files += files_in_space
            total_pages += pages_in_space
            
            # Debug mode: stop if reached limit
            if debug_mode and total_secrets >= debug_limit:
                logging.warning(f"🔍 DEBUG MODE: Reached {debug_limit} findings limit. Stopping scan and generating report...")
                break

    end_time = time.time()
    duration = end_time - start_time
    duration_str = f"{int(duration // 60)}m {int(duration % 60)}s"
    
    # Debug mode summary
    if debug_mode:
        logging.warning("=" * 80)
        logging.warning(f"🔍 DEBUG MODE: Stopped after {total_secrets} findings (limit: {debug_limit})")
        logging.warning("=" * 80)
    
    global_summary = f"[GLOBAL SUMMARY] Total scan duration: {duration_str}, {total_pages} pages scanned, {total_secrets} secrets found"
    if include_attachments:
        global_summary += f", {total_files} files scanned"
    logging.info(global_summary)
    
    logging.info(f"Results saved in: {output_file}")
    
    # Create XLSX report with appropriate name
    if include_attachments:
        xlsx_file = "confluence_secrets_in_files.xlsx"
    else:
        xlsx_file = "confluence_secrets.xlsx"
    
    xlsx_created = False
    if create_xlsx_report(output_file, xlsx_file):
        logging.info(f"XLSX report created: {xlsx_file}")
        xlsx_created = True
    else:
        logging.warning("XLSX report creation skipped or failed")
    
    # Send email if configured
    if email_sender and email_recipient:
        if not aws_region:
            aws_region = "eu-central-1"  # default
        
        logging.info("Preparing to send email report...")
        
        # Count unique affected spaces and pages from CSV
        affected_spaces_count, affected_pages_count = count_affected_from_csv(output_file)
        
        # Generate email body
        email_body = generate_email_summary(
            total_secrets=total_secrets,
            total_pages=total_pages,
            total_files=total_files,
            spaces_scanned=len(spaces),
            duration_str=duration_str,
            include_attachments=include_attachments,
            affected_spaces=affected_spaces_count,
            affected_pages=affected_pages_count
        )
        
        # Prepare subject
        if total_secrets > 0:
            subject = f"CRITICAL: Confluence Secrets Scanner - {total_secrets} Secrets Found"
        else:
            subject = "Confluence Secrets Scanner - No Secrets Found"
        
        # Send email with attachment
        attachment = xlsx_file if xlsx_created and os.path.exists(xlsx_file) else None
        
        send_email_with_attachment(
            subject=subject,
            body_text=email_body,
            sender=email_sender,
            recipient=email_recipient,
            aws_region=aws_region,
            attachment_path=attachment
        )
    else:
        logging.info("Email notification disabled (no sender/recipient configured)")
    
    # Send individual author alerts if enabled
    if alert and total_secrets > 0:
        if not email_sender:
            logging.warning("\n⚠️  To send alerts, --email-sender must be specified")
        else:
            logging.info("\n📨 Sending individual alerts to page editors...")
            
            # Parse CSV to group findings by editor email
            editors_findings = {}
            
            try:
                with open(output_file, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    header = next(reader)  # Skip header
                    
                    # Skip dummy row if present
                    first_row = next(reader, None)
                    if first_row and "DUMMY_ROW_DELETE_ME" in first_row[0]:
                        pass  # Already skipped
                    else:
                        # Process first row if it's not dummy
                        if first_row:
                            f.seek(0)  # Reset
                            next(reader)  # Skip header again
                    
                    for row in reader:
                        if not row or "DUMMY_ROW_DELETE_ME" in row[0]:
                            continue
                        
                        if include_attachments:
                            # With attachments: Space Name, Page Title, File Title, File Extension, File URL, Page URL, Keyword, Type, Value, Email
                            if len(row) >= 10:
                                editor_email = row[9].strip()
                                editor_name = editor_email  # Use email as name since we don't have display name
                                
                                finding = {
                                    'space_name': row[0],
                                    'page_title': row[1],
                                    'file_title': row[2],
                                    'file_extension': row[3],
                                    'file_url': row[4],
                                    'page_url': row[5],
                                    'keyword': row[6],
                                    'finding_type': row[7],
                                    'matched_value': row[8],
                                    'editor_email': editor_email,
                                    'is_attachment': True
                                }
                        else:
                            # Without attachments: Space Name, Page Title, Page URL, Keyword, Type, Value, Email
                            if len(row) >= 7:
                                editor_email = row[6].strip()
                                editor_name = editor_email  # Use email as name
                                
                                finding = {
                                    'space_name': row[0],
                                    'page_title': row[1],
                                    'page_url': row[2],
                                    'keyword': row[3],
                                    'finding_type': row[4],
                                    'matched_value': row[5],
                                    'editor_email': editor_email,
                                    'is_attachment': False
                                }
                        
                        # Skip if no valid email
                        if not editor_email or editor_email == 'N/A' or '@' not in editor_email or 'unknown' in editor_email:
                            continue
                        
                        if editor_email not in editors_findings:
                            editors_findings[editor_email] = {
                                'name': editor_name,
                                'findings': []
                            }
                        editors_findings[editor_email]['findings'].append(finding)
                
                # Send alerts to each editor
                logging.info(f"   Found {len(editors_findings)} editor(s) with secrets")
                
                success_count = 0
                for editor_email, data in editors_findings.items():
                    editor_name = data['name']
                    editor_findings = data['findings']
                    
                    logging.info(f"   📧 {editor_name} ({editor_email}): {len(editor_findings)} secret(s)")
                    
                    if send_author_alert(
                        author_email=editor_email,
                        author_name=editor_name,
                        author_findings=editor_findings,
                        aws_region=aws_region,
                        sender_email=email_sender,
                        security_contact=security_contact,
                        security_wiki=security_wiki,
                        has_attachments=include_attachments
                    ):
                        success_count += 1
                
                logging.info(f"\n   ✅ Successfully sent: {success_count}/{len(editors_findings)}")
                
            except Exception as e:
                logging.error(f"Error processing alerts: {e}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Improved Confluence Scanner v5 - Scan for secrets and keywords with enhanced regex support and XLSX output",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic scan with regex file
  python3 confluence_improved_v5.py --base-url https://your-org.atlassian.net \\
    --username user@example.com --token YOUR_TOKEN --regex-file regex.txt

  # Scan with attachments
  python3 confluence_improved_v5.py --base-url https://your-org.atlassian.net \\
    --username user@example.com --token YOUR_TOKEN --regex-file regex.txt \\
    --include-attachments --filetype docx,pdf,json

  # Public-only scan
  python3 confluence_improved_v5.py --base-url https://your-org.atlassian.net \\
    --public-only --regex-file regex.txt
    
  # Scan with author alerts
  python3 confluence_improved_v5.py --base-url https://your-org.atlassian.net \\
    --username user@example.com --token YOUR_TOKEN --regex-file regex.txt \\
    --email-sender security@company.com --alert --security-contact appsec@company.com
    
  # Send results to multiple recipients
  python3 confluence_improved_v5.py --base-url https://your-org.atlassian.net \\
    --username user@example.com --token YOUR_TOKEN --regex-file regex.txt \\
    --email-sender security@company.com --email-recipient "appsec@company.com,team@company.com"
        """
    )
    parser.add_argument("--base-url", help="Base Confluence URL")
    parser.add_argument("--username", help="API login (email) - not required with --public-only")
    parser.add_argument("--token", help="API token - not required with --public-only")
    parser.add_argument("--public-only", action="store_true", help="Scan only public spaces without authentication")
    parser.add_argument("--keywords", help="File with keywords (one per line)")
    parser.add_argument("--regex-file", help="Regex file with patterns in 'Name:::Regex:::GroupIndex' format")
    parser.add_argument("--regex", help="Single regex pattern (legacy)")
    parser.add_argument("--include-attachments", action="store_true", help="Scan attachment files")
    parser.add_argument("--filetype", help="Comma-separated file types to scan (only with --include-attachments), e.g., docx,pdf,json")
    parser.add_argument("--exclude-filetype", help="Comma-separated file types to exclude (only with --include-attachments), e.g., pdf")
    parser.add_argument("--max-size", help="Max file size to analyze (only with --include-attachments), e.g., 2mb, 500kb")
    parser.add_argument("--space-keys", help="Space keys to scan (comma-separated or file path)")
    parser.add_argument("--exclude-space-keys", help="Space keys to exclude (comma-separated or file path)")
    parser.add_argument("--modified-after", help="Scan pages modified after this date (D.M.Y or D/M/Y)")
    parser.add_argument("--modified-before", help="Scan pages modified before this date (D.M.Y or D/M/Y)")
    parser.add_argument("--created-in-year", help="Scan pages created in this year (e.g., 2025)")
    parser.add_argument("--no-duplicates", action="store_true", help="Exclude duplicate findings")
    parser.add_argument("--resume-from", help="Resume scanning from this space key")
    parser.add_argument("--max-attachment-age", help="Scan only recent attachments, e.g., 1d, 1w, 1m, 1y")
    parser.add_argument("--scan-images-only", action="store_true", help="Scan only images with OCR")
    parser.add_argument("--archive-support", action="store_true", help="Unpack and scan archives (zip, tar, etc.)")
    parser.add_argument("--secret-max-length", type=int, default=20, help="Maximum characters to display in 'Matched Value' column (default: 20)")
    parser.add_argument("--config", help="JSON config file with arguments")
    parser.add_argument("--debug", action="store_true", help="Debug mode: stop after first 5 findings and generate report")
    
    # Email notification arguments
    parser.add_argument("--email-sender", help="Sender email address for notifications (must be verified in AWS SES)")
    parser.add_argument("--email-recipient", help="Recipient email address(es) for scan results (single email or comma-separated list, e.g., 'user1@example.com,user2@example.com')")
    parser.add_argument("--aws-region", default="eu-central-1", help="AWS region for SES (default: eu-central-1)")
    
    # Author alert arguments
    parser.add_argument("--alert", action="store_true", help="Send individual email alerts to page editors who leaked secrets")
    parser.add_argument("--security-contact", default="security@company.com", help="Security team contact email for author alerts (default: security@company.com)")
    parser.add_argument("--security-wiki", help="Security documentation URL (optional, included in author alerts)")
    
    # Output file (moved to end)
    parser.add_argument("-o", "--output", default="confluence_results.csv", help="Output CSV file (XLSX will be auto-generated)")
    
    args = parser.parse_args()
    
    validation_errors = validate_arguments(args)
    if validation_errors:
        setup_logging()
        logging.error("Argument validation errors:")
        for error in validation_errors:
            logging.error(f"  - {error}")
        logging.info("Use --help for more information.")
        exit(1)

    main(
        base_url=args.base_url,
        username=args.username,
        token=args.token,
        keywords_file=args.keywords,
        regex_file=args.regex_file,
        single_regex=args.regex,
        output_file=args.output,
        include_attachments=args.include_attachments,
        filetypes=args.filetype,
        exclude_filetypes=args.exclude_filetype,
        max_size=args.max_size,
        public_only=args.public_only,
        space_keys=args.space_keys,
        exclude_space_keys=args.exclude_space_keys,
        modified_after=args.modified_after,
        modified_before=args.modified_before,
        created_in_year=args.created_in_year,
        no_duplicates=args.no_duplicates,
        resume_from=args.resume_from,
        max_attachment_age=args.max_attachment_age,
        scan_images_only=args.scan_images_only,
        archive_support=args.archive_support,
        config=args.config,
        email_sender=args.email_sender,
        email_recipient=args.email_recipient,
        aws_region=args.aws_region,
        secret_max_length=args.secret_max_length,
        debug_mode=args.debug,
        alert=args.alert,
        security_contact=args.security_contact,
        security_wiki=args.security_wiki
    )
